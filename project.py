from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from datetime import datetime, timedelta
import calendar

def create_habit_tracker():
    wb = Workbook()
    
    # Удаляем стандартный лист
    wb.remove(wb.active)
    
    # Список привычек с иконками
    habits_data = [
        ("🧘 Медитация", 10),
        ("💧 Пить воду", 31),
        ("🚶 10 000 шагов", 31),
        ("📖 Читать книгу", 10),
        ("🥗 Полезное питание", 31),
        ("🏋️ Тренировка", 10),
        ("😴 Сон до 23:00", 31),
        ("📓 Ведение дневника", 4),
        ("🧴 Уход за собой", 15),
        ("📵 Час без телефона", 31),
        ("📥 Проверить входящие заявки", 31),
        ("📱 Проверить соцсети", 31),
        ("📸 Выложить пост", 31),
        ("🎬 Снять/смонтировать сторис", 15),
        ("💬 Ответить клиентам", 31),
        ("📂 Разобрать рабочие чаты", 31),
        ("✅ Закрыть задачи в CRM", 31),
        ("📊 Проверить рекламу / статистику", 2),
        ("📄 Выставить счета / КП", 31),
        ("🗂 Навести порядок в проектах", 1),
    ]
    
    # Создаем лист Данные
    ws_data = wb.create_sheet("Данные")
    ws_data['A1'] = "Год"
    ws_data['B1'] = "Ваши привычки"
    ws_data['C1'] = "Сочетания клавиш для открытия панели эмодзи:"
    ws_data['A2'] = 2026
    ws_data['C2'] = "Windows: нажмите Windows + ; (точка с запятой) или Windows + . (точка)"
    ws_data['C3'] = "Mac: нажмите Control + Command + Пробел"
    
    for idx, (habit, goal) in enumerate(habits_data, start=2):
        ws_data[f'B{idx}'] = habit
    
    # Стили    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Месяцы на русском
    months_ru = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    
    weekdays_short = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    weekdays_full = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    
    # Создаем 12 листов для месяцев
    for month_idx in range(12):
        month_name = months_ru[month_idx]
        ws = wb.create_sheet(month_name)
        
        # Заголовок
        ws['A2'] = "ТРЕКЕР ПРИВЫЧЕК"
        ws['A2'].font = Font(bold=True, size=14)
        
        month_year = f"{month_name.upper()} 2026"
        ws['A3'] = month_year
        ws['A3'].font = Font(italic=True, size=12)
        
        # Календарь в правом верхнем углу (начиная с колонки AI)
        cal_start_col = 61  # AI
        
        # Получаем информацию о месяце
        cal = calendar.monthcalendar(2026, month_idx + 1)
        month_days = calendar.monthrange(2026, month_idx + 1)[1]
        
        # Заголовок календаря
        ws.merge_cells(start_row=2, start_column=cal_start_col, end_row=2, end_column=cal_start_col+6)
        ws.cell(row=2, column=cal_start_col, value=month_year).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=cal_start_col).font = Font(bold=True)
        
        # Дни недели
        for col, day_name in enumerate(weekdays_short, start=cal_start_col):
            ws.cell(row=3, column=col, value=day_name)
            ws.cell(row=3, column=col).font = Font(bold=True, size=8)
            ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')
        
        # Дни месяца
        row_offset = 4        for week in cal:
            for col, day in enumerate(week, start=cal_start_col):
                if day != 0:
                    ws.cell(row=row_offset, column=col, value=day)
                    ws.cell(row=row_offset, column=col).alignment = Alignment(horizontal='center')
            row_offset += 1
        
        # Заголовки недель
        ws.merge_cells(start_row=11, start_column=3, end_column=9)
        ws.merge_cells(start_row=11, start_column=10, end_column=16)
        ws.merge_cells(start_row=11, start_column=17, end_column=23)
        ws.merge_cells(start_row=11, start_column=24, end_column=30)
        ws.merge_cells(start_row=11, start_column=31, end_column=37)
        
        week_headers = ["1 неделя", "2 неделя", "3 неделя", "4 неделя", "5 неделя"]
        week_colors = ["FFE4D6", "#D6E4FF", "#E0F0E0", "#E8E0F0", "#FFE0E8"]
        
        for i, (header, color) in enumerate(zip(week_headers, week_colors)):
            cell = ws.cell(row=11, column=3 + i*7)
            cell.value = header
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Номера дней
        for day in range(1, 32):
            col = 2 + day
            ws.cell(row=12, column=col, value=day)
            ws.cell(row=12, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=12, column=col).font = Font(size=8)
            
            # День недели
            date = datetime(2026, month_idx + 1, day) if day <= month_days else None
            if date:
                weekday = weekdays_full[date.weekday()]
                ws.cell(row=13, column=col, value=weekday)
                ws.cell(row=13, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=13, column=col).font = Font(size=7)
        
        # Заголовки столбцов
        ws['B12'] = "ЦЕЛЬ"
        ws['B12'].font = Font(bold=True)
        
        # Заголовки справа
        headers_right = ["ИТОГО", "СЕРИЯ", "ВЫПОЛНЕНО", "ПРОГРЕСС"]
        for col, header in enumerate(headers_right, start=42):  # AJ
            ws.cell(row=12, column=col, value=header)
            ws.cell(row=12, column=col).font = Font(bold=True)
            ws.cell(row=12, column=col).alignment = Alignment(horizontal='center')
        
        # Добавляем привычки        for row_idx, (habit, goal) in enumerate(habits_data[:10], start=14):
            ws.cell(row=row_idx, column=1, value=habit)
            ws.cell(row=row_idx, column=2, value=goal)
            
            # Ячейки для отметок (дни месяца)
            for day in range(1, 32):
                col = 2 + day
                if day <= month_days:
                    ws.cell(row=row_idx, column=col, value="")
            
            # Формулы для подсчета
            total_col = 42  # AJ
            streak_col = 43  # AK
            completed_col = 44  # AL
            progress_col = 45  # AM
            
            # ИТОГО - считаем количество X
            formula_total = f'=COUNTIF(C{row_idx}:AG{row_idx},"x")'
            ws.cell(row=row_idx, column=total_col, value=formula_total)
            
            # СЕРИЯ (максимальная последовательность)
            ws.cell(row=row_idx, column=streak_col, value=0)
            
            # ВЫПОЛНЕНО
            formula_completed = f'=AJ{row_idx}/{goal}'
            ws.cell(row=row_idx, column=completed_col, value=formula_completed)
            
            # ПРОГРЕСС в процентах
            formula_progress = f'=AL{row_idx}*100'
            ws.cell(row=row_idx, column=progress_col, value=formula_progress)
            ws.cell(row=row_idx, column=progress_col).number_format = '0%'
        
        # Применяем границы и форматирование
        for row in range(11, 25):
            for col in range(1, 46):
                cell = ws.cell(row=row, column=col)
                if row == 11 or row == 12 or row == 13 or col == 1 or col == 2 or col >= 42:
                    cell.border = Border(bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Настраиваем ширину колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 8
        for day in range(1, 32):
            col_letter = get_column_letter(day + 2)
            ws.column_dimensions[col_letter].width = 3
        
        # Скрываем ненужные колонки после 31 дня
        for col in range(34, 42):
            col_letter = get_column_letter(col)            ws.column_dimensions[col_letter].width = 10
        
        ws.column_dimensions['AJ'].width = 8
        ws.column_dimensions['AK'].width = 8
        ws.column_dimensions['AL'].width = 12
        ws.column_dimensions['AM'].width = 12
        
        # Добавляем прогресс-бары
        for row_idx in range(14, 24):
            progress_cell = f'AM{row_idx}'
            ws.conditional_formatting.add(
                progress_cell,
                DataBarRule(
                    start_type='num',
                    start_value=0,
                    end_type='num',
                    end_value=100,
                    color="00B050",
                    showValue=True
                )
            )
        
        # Добавляем формулу для общего прогресса месяца
        ws['A4'] = f"=TEXT(AVERAGE(AM14:AM23), \"0%\")"
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
    
    # Удаляем последний лишний лист (если есть)
    if len(wb.sheetnames) > 13:
        wb.remove(wb.sheetnames[-1])
    
    # Сохраняем
    wb.save('Трекер_привычек_2026.xlsx')
    print("✅ Трекер привычек успешно создан: Трекер_привычек_2026.xlsx")

if __name__ == "__main__":
    create_habit_tracker()